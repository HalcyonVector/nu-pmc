from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ─── Styling ───
NAVY = '1D3D62'
BG_HEADER = PatternFill('solid', start_color=NAVY)
BG_SUBHEADER = PatternFill('solid', start_color='E8EEF4')
BG_YELLOW = PatternFill('solid', start_color='FFFBEF')
BG_GREY = PatternFill('solid', start_color='F8F7F3')

F_TITLE = Font(name='Arial', size=14, bold=True, color=NAVY)
F_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY = Font(name='Arial', size=10)
F_BODY_BOLD = Font(name='Arial', size=10, bold=True)
F_MUTED = Font(name='Arial', size=9, color='888888', italic=True)

thin = Side(border_style='thin', color='CCCCCC')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ═══════════════════════════════════════════════
# ROLE_TABS — authoritative source (from app.js lines 13-36)
# ═══════════════════════════════════════════════
ROLE_TABS = {
    'Principal': ['dashboard','monthly','projects','project_detail','approvals','signoff','changes','delegations','budget','budget_tree','vendors','vendors_master','boq_mapping','materials','client_boq','payments','payments_fin','finance_clearance','weekly_health','users'],
    'Design Principal': ['dashboard','monthly','projects','project_detail','approvals','signoff','changes','delegations','budget','vendors','vendors_master','boq_mapping','materials','client_boq','payments','payments_fin','finance_clearance','weekly_health','users'],
    'PMC Head': ['dashboard','monthly','project_detail','reports','signoff','grn','issues','meetings','delegations','vendors','vendors_master','boq_mapping','materials','client_boq','payments','labour','deputy'],
    'Design Head': ['dashboard','monthly','project_detail','drawings','register','signoff','issues','submittals','delegations','phototags','vendors','vendors_master','boq_mapping','materials','client_boq','payments','budget'],
    'Services Head': ['dashboard','monthly','project_detail','drawings','register','signoff','issues','submittals','delegations','phototags','vendors','vendors_master','boq_mapping','materials','client_boq','payments','budget'],
    'Site Manager': ['dashboard','project_detail','tasks','drawings','register','grn','issues','issues_site','labour','photos','materials_site','payments'],
    'Senior Site Manager': ['dashboard','project_detail','tasks','drawings','register','grn','issues','issues_site','labour','photos','materials_site','payments'],
    'Finance Admin': ['dashboard','monthly','project_detail','clients','vendors_master','finance_clearance','payments_fin','pi','petty_cash','client_receipts','gst_statement','client_boq'],
    'Trainee': ['drawings','schedule_view'],
    'Team Lead': ['dashboard','project_detail','drawings','issues','submittals'],
    'Detailing Head': ['dashboard','project_detail','drawings','register','issues','submittals','phototags'],
    'Jr Architect': ['dashboard','project_detail','drawings','issues','submittals'],
    'Detailing': ['drawings','submittals'],
    'Services Engineer': ['dashboard','project_detail','drawings','issues','submittals','phototags'],
    'Coordinator': ['dashboard','project_detail','drawings','register','issues','tasks','grn','meetings','photos'],
    'Audit': ['dashboard','monthly','projects','project_detail','approvals','signoff','changes','delegations','budget','budget_tree','vendors','vendors_master','boq_mapping','materials','client_boq','payments','payments_fin','finance_clearance','weekly_health','users','clients','drawings','register','grn','issues','meetings','labour','deputy','phototags','issues_site','tasks','photos','submittals','pi','petty_cash','client_receipts','gst_statement','reports','schedule_view','ncr','compliance','tally','notifications'],
}

# ═══════════════════════════════════════════════
# TAB descriptions
# Each tab: { label, what_inside (default role-agnostic), per_role_override (dict) }
# per_role_override lets a specific role see different content on the same tab.
# ═══════════════════════════════════════════════
TABS = {
    'dashboard': {
        'label': 'Dashboard',
        'default': 'Action Centre — overdue queries, open site flags, overdue materials, pending approvals awaiting principals, fresh drawing queries within 3 days. Tap any item to jump to that tab.',
        'override': {
            'Site Manager': 'Action Centre — their overdue items, site flags awaiting their response, materials overdue on their project.',
            'Senior Site Manager': 'Action Centre — same as site manager, plus items across all their assigned projects.',
            'Finance Admin': 'Action Centre — pending payments ready for ICICI upload, rejected payments to follow up, GST filing due dates.',
            'Audit': 'Action Centre — everything across all projects, read-only.',
        }
    },
    'monthly': {
        'label': 'Monthly Overview',
        'default': 'Grid of every active project with trade-progress bars (Civil, Electrical, HVAC etc.), status badge, R0 end date, and 3 counters per project: open queries, flagged tasks, overdue materials.',
    },
    'projects': {
        'label': 'Projects',
        'default': 'Full project list — every project with phase, progress%, location, status. Tap to drill into Project Summary.',
    },
    'project_detail': {
        'label': 'Project Summary',
        'default': '⚠ Being redesigned — see Role × Button matrix (currently shows name/client/location/contract value + role-specific buttons).',
    },
    'approvals': {
        'label': 'Approvals',
        'default': 'All pending approvals in the queue — schedule changes, weekly reports, change notices, anything needing their sign-off. Tap to approve/reject individual items.',
        'override': {
            'Audit': 'Read-only view of every pending approval across all projects.',
        }
    },
    'signoff': {
        'label': 'Weekly Sign-off',
        'default': 'Weekly sign-off summary — open items needing their sign-off, approved items, rejected items. Workflow for accountability tracking.',
    },
    'changes': {
        'label': 'CNs',
        'default': 'Change Notices — list of active CNs, their ₹ impact, who approved, pending ones. Create new CN. Approve below ₹1L (peer), above ₹1L routes to principal.',
    },
    'delegations': {
        'label': 'Delegations',
        'default': 'Delegation log — who is standing in for whom while on leave. Grant temporary access to another user. View active + past delegations.',
    },
    'budget': {
        'label': 'Budget',
        'default': 'Project budget — cost heads, sanctioned amounts, spent-to-date, variance. Approve custom cost heads (principals + stream heads).',
        'override': {
            'Design Head': 'Budget tab restricted to design-stream cost heads.',
            'Services Head': 'Budget tab restricted to services-stream cost heads.',
        }
    },
    'budget_tree': {
        'label': 'Budget Tree',
        'default': 'Hierarchical variance view — cost head → sub-head → trade → line item. Variance coloring (green/amber/red). Drill-down tree.',
    },
    'vendors': {
        'label': 'Engagements',
        'default': 'Vendor engagements — active engagements per project, contract values, status (raised / approved / active). Raise new engagement, view contract revisions.',
    },
    'vendors_master': {
        'label': 'Vendor Master',
        'default': 'Master vendor database — every vendor with GST, PAN, bank details, clearance status, category. Register new vendor. Edit master (finance + principals).',
    },
    'boq_mapping': {
        'label': 'BOQ Map',
        'default': 'Map client BOQ line items to vendors + trades. Bulk assign. Shows unassigned items needing attention.',
    },
    'materials': {
        'label': 'Materials',
        'default': 'Material approvals master — specs, samples, brands, approval status. Approve/reject material submissions.',
        'override': {
            'Design Head': 'Design-stream materials only (finishes, fittings).',
            'Services Head': 'Services-stream materials only (HVAC, electrical, plumbing).',
        }
    },
    'materials_site': {
        'label': 'Materials',
        'default': 'Site view — materials approved for this project, expected delivery dates, overdue items, GRN status.',
    },
    'client_boq': {
        'label': 'Client BOQ',
        'default': 'Client-side Bill of Quantities — items, quantities, rates, totals. Version history. Upload/replace.',
    },
    'payments': {
        'label': 'Payments',
        'default': 'Payment requests list — raised, pending PMC/principal approval, approved, paid, rejected. Raise new PR.',
        'override': {
            'Site Manager': 'Payment requests they raised or are aware of on this project. Upload invoice/photo to raise.',
            'Senior Site Manager': 'Same as site manager, plus GRN-approval power up to 5% of project budget.',
        }
    },
    'payments_fin': {
        'label': 'Payments (Finance)',
        'default': 'Finance view — Naveen-approved PRs ready for ICICI bulk upload, download ICICI Excel, upload UTR confirmation file, payment history with UTRs.',
    },
    'finance_clearance': {
        'label': 'Vendor Clearance',
        'default': 'Finance queue of new vendors awaiting clearance (GST check, PAN validation, bank verification). Approve for payment eligibility or reject.',
    },
    'weekly_health': {
        'label': 'Health',
        'default': 'Weekly project health scores — schedule drift, budget variance, open flags, RFI overdue count, overall health indicator. Trend chart.',
    },
    'users': {
        'label': 'Users',
        'default': 'User management — list of all users, roles, is_active, last login. Create/deactivate. Reset passwords. Change roles (principals only).',
    },
    'clients': {
        'label': 'Clients',
        'default': 'Client master — firm name, GSTIN, PAN, state, billing address, tally ledger codes. Create/edit. Mark complete.',
    },
    'drawings': {
        'label': 'Drawings',
        'default': 'All drawings — filter by category, stream, status. Latest approved version per drawing, version history, PDF preview.',
        'override': {
            'Design Head': 'Design-stream drawings. Upload new, approve/reject versions. Flag drawings.',
            'Services Head': 'Services-stream drawings. Upload, approve/reject, flag.',
            'Site Manager': 'Read-only — view latest approved version of every drawing on the project. Tap to open PDF.',
            'Senior Site Manager': 'Read-only — view latest approved version of every drawing on the project. Tap to open PDF.',
            'Coordinator': 'Read-only — view all approved drawings. Cannot upload/approve.',
            'Trainee': 'Read-only — view approved drawings only (no drafts, no flagged).',
            'Team Lead': 'Design/services heads\' team — view + preview. Upload detail drawings or RFI responses for their team.',
            'Jr Architect': 'Upload detail drawings and RFI responses. Cannot approve/reject.',
            'Services Engineer': 'Same as jr_architect but services stream.',
            'Detailing Head': 'Upload detail drawings. Sign-off on register entries.',
            'Detailing': 'Upload detail drawings only. No register, no approvals.',
        }
    },
    'register': {
        'label': 'Register',
        'default': 'Drawing register — master list of every expected main drawing, status (pending/in-progress/issued), summary counts.',
        'override': {
            'Design Head': 'Design register. Upload/amend entries. Sign-off.',
            'Services Head': 'Services register. Upload/amend. Sign-off.',
            'Site Manager': 'Read-only view of entire register — knows which drawings are coming, their status.',
            'Senior Site Manager': 'Read-only view of register.',
            'Coordinator': 'Read-only.',
            'Detailing Head': 'View + sign-off rights but no amend.',
        }
    },
    'submittals': {
        'label': 'Submittals',
        'default': 'Submittal log — items awaiting review, review comments, approved, resubmit-required, rejected. Review submittals (heads + pmc + principals). Upload new (design team).',
        'override': {
            'Design Head': 'Review design-stream submittals. Approve / approve-with-comments / resubmit / reject.',
            'Services Head': 'Review services-stream submittals.',
            'Team Lead': 'Upload submittals for their team.',
            'Jr Architect': 'Upload submittals.',
            'Services Engineer': 'Upload submittals (services stream).',
            'Detailing Head': 'View + upload.',
            'Detailing': 'Upload detail submittals.',
        }
    },
    'grn': {
        'label': 'GRNs',
        'default': 'Goods Received Notes — raised, pending approval, approved, flagged for NCR. Raise new GRN. Approve (PMC head, senior site below 5% budget). Flag non-conformance.',
        'override': {
            'Site Manager': 'Raise new GRN when materials arrive. View own GRNs + status.',
            'Senior Site Manager': 'Raise GRN + approve below 5% of project budget.',
            'Coordinator': 'View GRN log, raise on behalf of site team.',
        }
    },
    'issues': {
        'label': 'Issues',
        'default': 'Issues log — RFI (request for info), design queries, safety, quality, compliance. Raise, assign, close.',
        'override': {
            'Design Head': 'Design-stream issues routed to them.',
            'Services Head': 'Services-stream issues.',
            'Site Manager': 'Field issues on this project — raise, respond to closures.',
            'Senior Site Manager': 'All field issues.',
            'Team Lead': 'Issues assigned to their team.',
            'Jr Architect': 'RFI and design queries assigned to them.',
            'Services Engineer': 'Services RFIs and queries.',
            'Coordinator': 'Cross-issue view for coordination tracking.',
        }
    },
    'issues_site': {
        'label': 'Field Issues',
        'default': 'Site-raised issues — safety flags, quality concerns, snags during work. Raise quickly with photo. Track status.',
    },
    'meetings': {
        'label': 'Meetings',
        'default': 'MOMs (Minutes of Meeting) and site visit reports. Draft MOM → PMC Head approves → issue to client. View action items, countersign assignments, mark completed.',
        'override': {
            'PMC Head': 'Approve draft MOMs (checkpoint before client issue). Issue to client (separate click).',
            'Coordinator': 'Draft MOMs, upload attachments, manage action items.',
        }
    },
    'labour': {
        'label': 'Labour',
        'default': 'Labour compliance tracker — contractor licenses, PF/ESI records, daily headcount, compliance flags.',
        'override': {
            'Site Manager': 'Daily headcount entry, compliance photo uploads.',
            'Senior Site Manager': 'Headcount entry plus compliance overview.',
        }
    },
    'deputy': {
        'label': 'Deputy',
        'default': 'PMC deputy management — assign deputy authority when PMC head is on leave. Override deputy decisions. Principal-level controls.',
    },
    'phototags': {
        'label': 'Photo Review',
        'default': 'Photo tagging review — site photos pending tagging, tagged photos for heads to review. Approve/reject tags. Surface issues found in photos.',
    },
    'photos': {
        'label': 'Photos',
        'default': 'Site photos — upload daily photos, tag them, browse by date, by location, by trade.',
        'override': {
            'Site Manager': 'Upload site photos (camera/gallery), tag location + trade.',
            'Senior Site Manager': 'Upload + review team photos.',
        }
    },
    'tasks': {
        'label': 'Tasks',
        'default': 'Schedule view — today/week/look-ahead tasks scheduled on this project. Mark % complete. Flag delays.',
        'override': {
            'Site Manager': 'Today\'s tasks assigned to them. Mark pct_complete, upload photo per task, flag delays.',
            'Senior Site Manager': 'Today + look-ahead tasks for their team.',
            'Coordinator': 'Schedule view for coordination.',
            'PMC Head': '(Not currently in their sidebar — they use reports + approvals instead.)',
        }
    },
    'schedule_view': {
        'label': 'Schedule',
        'default': 'Read-only schedule view — full project schedule, trades, milestones, drift. For roles that should see schedule but not edit.',
    },
    'reports': {
        'label': 'Reports',
        'default': 'Daily reports — pending review, AI-flagged anomalies, approved history. Batch approve (PMC only). Review individual reports. Flag for site team.',
    },
    'pi': {
        'label': 'Invoices',
        'default': 'Performa Invoices (client invoicing) — draft PI, generate Tally-compatible Excel, email client. History + status.',
    },
    'petty_cash': {
        'label': 'Petty Cash',
        'default': 'Petty cash ledger — running balance, transactions (spend/replenish), per-project tracking.',
    },
    'client_receipts': {
        'label': 'Receipts',
        'default': 'Client receipts — log incoming client payments, allocate to invoices/projects, update client ledger.',
    },
    'gst_statement': {
        'label': 'GST',
        'default': 'Monthly GST statement — auto-generated from invoices + receipts. Download Excel for CA. Shows output GST, input GST, net liability.',
    },
    'ncr': {
        'label': 'NCRs',
        'default': 'Non-Conformance Reports — from flagged GRNs or quality issues. Close-out tracking.',
    },
    'compliance': {
        'label': 'Compliance',
        'default': 'Compliance tracker — statutory deadlines, labour compliance expiry, approval expiry (BBM/BDA/ELCITA).',
    },
    'tally': {
        'label': 'Tally',
        'default': 'Tally export — vouchers, journals, purchase/sales entries in Tally-compatible format.',
    },
    'notifications': {
        'label': 'Alerts',
        'default': 'Alert log — WhatsApp-sent notifications history, in-app notifications, read/unread.',
    },
}

# ═══════════════════════════════════════════════
# BUILD — one sheet per role
# ═══════════════════════════════════════════════

# Add summary sheet first
ws_sum = wb.create_sheet('Summary')
ws_sum['A1'] = 'nu PMC — Per-Role Sidebar Audit'
ws_sum['A1'].font = F_TITLE
ws_sum.merge_cells('A1:C1')
ws_sum['A2'] = 'One sheet per role. Each sheet: tabs as columns, cell content describes what they see inside.'
ws_sum['A2'].font = F_MUTED
ws_sum.merge_cells('A2:C2')
ws_sum['A3'] = 'Yellow cells = role-specific behavior (different from default). Plain cells = same as default.'
ws_sum['A3'].font = F_MUTED
ws_sum.merge_cells('A3:C3')

hdrs = ['Role', 'Tab count', 'Notes']
for col_idx, h in enumerate(hdrs, 1):
    c = ws_sum.cell(row=5, column=col_idx, value=h)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = CENTER
    c.border = BOX

row = 6
for role, tabs in ROLE_TABS.items():
    c = ws_sum.cell(row=row, column=1, value=role)
    c.font = F_BODY_BOLD
    c.fill = BG_SUBHEADER
    c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='center')
    c2 = ws_sum.cell(row=row, column=2, value=len(tabs))
    c2.font = F_BODY
    c2.alignment = CENTER
    c2.border = BOX
    # Notes — flag if no project_detail tab
    note = []
    if 'project_detail' not in tabs:
        note.append('no Project Summary')
    if len(tabs) <= 3:
        note.append('minimal access')
    if role == 'Audit':
        note.append('sees everything (read-only)')
    c3 = ws_sum.cell(row=row, column=3, value=' · '.join(note) if note else '')
    c3.font = F_BODY
    c3.alignment = LEFT
    c3.border = BOX
    row += 1

ws_sum.column_dimensions['A'].width = 24
ws_sum.column_dimensions['B'].width = 12
ws_sum.column_dimensions['C'].width = 40

# ─── Per-role sheets ───
for role, tabs in ROLE_TABS.items():
    # Sheet name max 31 chars
    sheet_name = role[:31]
    ws = wb.create_sheet(sheet_name)
    ws['A1'] = f'{role} — what they see in each tab'
    ws['A1'].font = F_TITLE
    ws.merge_cells('A1:D1')
    ws['A2'] = f'{len(tabs)} tabs in sidebar. Yellow = role-specific content (overrides default).'
    ws['A2'].font = F_MUTED
    ws.merge_cells('A2:D2')

    # Header row — tab labels
    for col_idx, tab_key in enumerate(tabs, 1):
        tab_def = TABS.get(tab_key, {'label': tab_key, 'default': '(no description available)'})
        c = ws.cell(row=4, column=col_idx, value=tab_def['label'])
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = CENTER
        c.border = BOX
    ws.row_dimensions[4].height = 32

    # Content row
    for col_idx, tab_key in enumerate(tabs, 1):
        tab_def = TABS.get(tab_key, {'label': tab_key, 'default': '(no description available)'})
        override = tab_def.get('override', {}).get(role)
        if override:
            content = override
            fill = BG_YELLOW
        else:
            content = tab_def['default']
            fill = None
        c = ws.cell(row=5, column=col_idx, value=content)
        c.font = F_BODY
        c.alignment = LEFT
        c.border = BOX
        if fill:
            c.fill = fill
    # Make content row tall enough
    ws.row_dimensions[5].height = 280

    # Set column widths
    for i in range(1, len(tabs) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 32

    # Notes row
    ws.cell(row=7, column=1, value='Your audit notes').font = F_BODY_BOLD
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max(4, len(tabs)))
    for col_idx in range(1, len(tabs) + 1):
        c = ws.cell(row=8, column=col_idx, value='')
        c.fill = BG_YELLOW
        c.border = BOX
        c.alignment = LEFT
    ws.row_dimensions[8].height = 80

# Save
import os
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
OUT = '/mnt/user-data/outputs/20260420 nu RoleSidebar Audit v1.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheets: {len(wb.sheetnames)} (1 summary + {len(ROLE_TABS)} role sheets)')
