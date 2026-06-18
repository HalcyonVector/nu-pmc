from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

NAVY   = '1D3D62'
BG_NAV  = PatternFill('solid', start_color=NAVY)
BG_HOME  = PatternFill('solid', start_color='D6E4F5')
BG_WORK  = PatternFill('solid', start_color='FFF0D6')
BG_MONEY = PatternFill('solid', start_color='D6EFE0')
BG_MORE  = PatternFill('solid', start_color='EDE8DF')
BG_PEND  = PatternFill('solid', start_color='F5D6D6')
BG_STRIP = PatternFill('solid', start_color='E8E4DC')
BG_HIDE  = PatternFill('solid', start_color='F8F8F8')
BG_SUB   = PatternFill('solid', start_color='EEF3F9')
BG_YEL   = PatternFill('solid', start_color='FFFBEF')
BG_FLAG  = PatternFill('solid', start_color='FFF3CD')

BFILL = {'Home':BG_HOME,'Work':BG_WORK,'Money':BG_MONEY,'More':BG_MORE,
         'Pending':BG_PEND,'—':BG_HIDE,'2-tab strip':BG_STRIP}

F_T = Font(name='Arial', size=14, bold=True, color=NAVY)
F_H = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_S = Font(name='Arial', size=10, bold=True, color=NAVY)
F_B = Font(name='Arial', size=10)
F_M = Font(name='Arial', size=9, color='888888', italic=True)
F_F = Font(name='Arial', size=9, bold=True, color='8B4513')

thin = Side(border_style='thin', color='CCCCCC')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
LC= Alignment(horizontal='left',   vertical='center', wrap_text=True)

def hdr(ws, r, c, val, w=None):
    x = ws.cell(row=r, column=c, value=val)
    x.font=F_H; x.fill=BG_NAV; x.alignment=C; x.border=BOX
    if w: ws.column_dimensions[get_column_letter(c)].width=w
def cel(ws, r, c, val, fill=BG_HIDE, font=None, align=None):
    x = ws.cell(row=r, column=c, value=val)
    x.font=font or F_B; x.fill=fill; x.alignment=align or L; x.border=BOX
    return x

# ═══════════════════════════════
# SHEET 1 — MASTER NAV TABLE
# ═══════════════════════════════
ws = wb.create_sheet('1 · Master Nav')
ws['A1'] = 'nu PMC — Navigation Redesign · All Roles'; ws['A1'].font=F_T; ws.merge_cells('A1:G1')
ws['A2'] = 'Bottom bar: max 4 buttons. Empty buckets hide. Locked decisions — sign off before build.'; ws['A2'].font=F_M; ws.merge_cells('A2:G2')

for c,(v,w) in enumerate([('Role',24),('Btn 1',14),('Btn 2',14),('Btn 3',14),('Btn 4',14),('Pinned section in Work',32),('Notes',36)],1):
    hdr(ws,4,c,v,w)
ws.row_dimensions[4].height=28

ROLES=[
    ('Principal',          'Home','Money','Pending','More',  'None (no Work bucket)','Pending=Blocked+Needs You. Drawings via Project Summary only.'),
    ('Design Principal',   'Home','Money','Pending','More',  'None (no Work bucket)','Same as Principal'),
    ('PMC Head',           'Home','Work', 'Money',  'Pending','⚡ Needs You (reports/MOMs/PRs/GRNs to action)','Pending=Blocked only (site team overdue)'),
    ('Design Head',        'Home','Work', 'Money',  'More',  '⚡ Needs You (drawings/submittals/materials pending)','Payments=advance+final only'),
    ('Services Head',      'Home','Work', 'Money',  'More',  '⚡ Needs You (services stream)','Same structure as Design Head'),
    ('Site Manager',       'Home','Work', 'Money',  '—',     '📋 Today\'s Report (submit by 6pm)','No 4th button'),
    ('Senior Site Manager','Home','Work', 'Money',  '—',     '⚡ Needs You (GRN approvals)\n📋 Today\'s Report','Two pinned items in Work'),
    ('Finance Admin',      'Home','Money','—',       '—',     '—','Only 2 buttons. Udupa knows his tools.'),
    ('Team Lead',          'Home','Work', '—',       '—',     '⚡ Team Needs You (team\'s pending items)','Merged with Detailing Head (same person)'),
    ('Jr Architect',       'Home','Work', '—',       '—',     '⚡ My Needs You (own items only)',''),
    ('Services Engineer',  'Home','Work', '—',       '—',     '⚡ My Needs You (own items only)',''),
    ('Coordinator',        'Home','Work', '—',       '—',     '⚡ Needs You (MOM actions, issues assigned)',''),
    ('Trainee',            '2-tab strip','—','—',    '—',     '—','No bottom bar. Top strip: Drawings · Schedule'),
    ('Detailing',          '2-tab strip','—','—',    '—',     '—','No bottom bar. Top strip: Drawings · Submittals'),
    ('Audit',              'Home','Money','Pending', 'More',  '—','Read-only mirror of Principal. No action buttons.'),
]

r=5
for role,b1,b2,b3,b4,pinned,note in ROLES:
    cel(ws,r,1,role,BG_SUB,F_S,LC)
    for c,v in [(2,b1),(3,b2),(4,b3),(5,b4)]:
        cel(ws,r,c,v,BFILL.get(v,BG_HIDE),F_B,C)
    cel(ws,r,6,pinned,BG_YEL if pinned!='—' else BG_HIDE,F_B,L)
    cel(ws,r,7,note,BG_HIDE,Font(name='Arial',size=9,color='555555'),L)
    ws.row_dimensions[r].height=36
    r+=1

r+=1
cel(ws,r,1,'Your audit notes:',BG_SUB,F_S,LC); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
r+=1
for c in range(1,8): cel(ws,r,c,'',BG_YEL)
ws.row_dimensions[r].height=60

# ═══════════════════════════════
# SHEET 2 — BUCKET CONTENTS
# ═══════════════════════════════
ws2 = wb.create_sheet('2 · Bucket Contents')
ws2['A1']='nu PMC — What\'s Inside Each Bucket · Frequency Ordered'; ws2['A1'].font=F_T; ws2.merge_cells('A1:F1')
ws2['A2']='Tabs listed most→least frequent. Sort toggle on all lists: Urgency / Age / Default.'; ws2['A2'].font=F_M; ws2.merge_cells('A2:F2')

for c,(v,w) in enumerate([('Role',20),('Home',22),('Work',34),('Money',28),('More',24),('Pending',30)],1):
    hdr(ws2,4,c,v,w)
ws2.row_dimensions[4].height=28

CONTENTS=[
    ('Principal',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary\n(drawings via Project Summary)',
     '—',
     'Payments\nVendors (Master+Clearance)\nBudget (Summary+Tree)\nVendor Allocation\nClient Contract',
     'Delegations\nCN Register\nHealth Dashboard\nUsers',
     '🔴 Blocked:\n  Drawings overdue w/ Design Head\n  RFIs overdue\n  GRNs overdue w/ PMC\n  MOMs not approved\n  Vendor clearance overdue\n  Health gone red/amber\n📋 Needs You:\n  PRs naveen-review\n  Engagements to approve\n  Client claims\n  CNs above ₹1L\n  Sign-off pending'),

    ('Design Principal',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary',
     '—',
     'Payments\nVendors\nBudget\nVendor Allocation\nClient Contract',
     'Delegations\nCN Register\nHealth Dashboard\nUsers',
     'Same as Principal'),

    ('PMC Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  Reports to approve N\n  MOMs awaiting approval N\n  PRs to PMC-review N\n  GRNs to sign off N\n──────────\nReports\nIssues\nMeetings\nDrawings\nRegister\nMaterials\nLabour',
     'GRNs\nPayments\nVendors\nVendor Master',
     '—',
     '🔴 Blocked only:\n  Reports not submitted by site\n  Issues no update from site\n  Labour not entered\n  Photos not uploaded'),

    ('Design Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  Drawings pending review N\n  Submittals to review N\n  Material approvals N\n──────────\nDrawings\nIssues\nSubmittals\nRegister\nPhoto Review',
     'Materials\nBudget\nPayments (advance+final only)',
     'Weekly Sign-off\nDelegations',
     '—'),

    ('Services Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned)\n[services stream]\n──────────\nDrawings\nIssues\nSubmittals\nRegister\nPhoto Review',
     'Materials\nBudget\nPayments (advance+final only)',
     'Weekly Sign-off\nDelegations',
     '—'),

    ('Site Manager',
     'Dashboard\nProject Summary',
     '📋 TODAY\'S REPORT (pinned)\n  [submit by 6pm]\n  Status: Not submitted / Submitted / Published\n──────────\nTasks\nPhotos\nField Issues\nIssues\nLabour\nDrawings\nRegister',
     'GRNs\nPayments\nMaterials',
     '—',
     '—'),

    ('Senior Site Manager',
     'Dashboard\nProject Summary',
     '⚡ NEEDS YOU (pinned)\n  GRN approvals (below 5% budget)\n📋 TODAY\'S REPORT (pinned)\n──────────\nTasks\nPhotos\nField Issues\nIssues\nLabour\nDrawings\nRegister',
     'GRNs\nPayments\nMaterials',
     '—',
     '—'),

    ('Finance Admin',
     'Dashboard\nMonthly Overview\nProject Summary',
     '—',
     'Payments (Finance)\nVendors (Master+Clearance)\nClient Receipts\nPetty Cash\nInvoices (PI)\nGST\nClient Contract\nClients',
     '—',
     '—'),

    ('Team Lead',
     'Dashboard\nProject Summary',
     '⚡ TEAM NEEDS YOU (pinned)\n  Team\'s pending RFIs N\n  Submittals returned N\n──────────\nDrawings\nIssues\nSubmittals\nRegister',
     '—','—','—'),

    ('Jr Architect',
     'Dashboard\nProject Summary',
     '⚡ MY NEEDS YOU (pinned)\n  Own RFIs assigned N\n  Submittals returned N\n──────────\nDrawings\nIssues\nSubmittals',
     '—','—','—'),

    ('Services Engineer',
     'Dashboard\nProject Summary',
     '⚡ MY NEEDS YOU (pinned)\n──────────\nDrawings\nIssues\nSubmittals\nPhoto Review',
     '—','—','—'),

    ('Coordinator',
     'Dashboard\nProject Summary',
     '⚡ NEEDS YOU (pinned)\n  MOM action items N\n  Issues assigned N\n──────────\nMeetings\nTasks\nIssues\nDrawings\nRegister\nPhotos\nGRNs',
     '—','—','—'),

    ('Trainee','—',
     '2-TAB STRIP (no bottom bar)\nDrawings  |  Schedule',
     '—','—','—'),

    ('Detailing','—',
     '2-TAB STRIP (no bottom bar)\nDrawings  |  Submittals',
     '—','—','—'),

    ('Audit',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary',
     'All tabs — read only',
     'All tabs — read only',
     'All tabs — read only',
     'All pending items — read only'),
]

r=5
for row_data in CONTENTS:
    role=row_data[0]
    cel(ws2,r,1,role,BG_SUB,F_S,LC)
    fills=[BG_HOME,BG_WORK,BG_MONEY,BG_MORE,BG_PEND]
    for c,val in enumerate(row_data[1:],2):
        cel(ws2,r,c,val,fills[c-2] if val!='—' else BG_HIDE)
    ws2.row_dimensions[r].height=200 if role not in ('Trainee','Detailing','Finance Admin') else 60
    r+=1

r+=1
cel(ws2,r,1,'Your audit notes:',BG_SUB,F_S,LC); ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
r+=1
for c in range(1,7): cel(ws2,r,c,'',BG_YEL)
ws2.row_dimensions[r].height=80

# ═══════════════════════════════
# SHEET 3 — CONSOLIDATIONS
# ═══════════════════════════════
ws3 = wb.create_sheet('3 · Consolidations')
ws3['A1']='nu PMC — Tab Consolidations & Renames Locked'; ws3['A1'].font=F_T; ws3.merge_cells('A1:D1')
ws3['A2']='These reduce tab count and clarify naming. All locked — confirm before build.'; ws3['A2'].font=F_M; ws3.merge_cells('A2:D2')

for c,(v,w) in enumerate([('Change',24),('Old',24),('New',24),('Notes',36)],1):
    hdr(ws3,4,c,v,w)

CONS=[
    ('Merge',      'Vendor Master + Vendor Clearance','One "Vendors" tab','Clearance is a filtered section inside Master. Finance Admin and principals see both in one place.'),
    ('Merge',      'Budget + Budget Tree',            'One "Budget" tab (toggle inside)','Summary view and drill-down tree as a toggle. Same data, two presentations.'),
    ('Rename',     'BOQ Map',                         'Vendor Allocation','Clearer — it\'s about which vendor covers which BOQ line item'),
    ('Rename',     'Client BOQ',                      'Client Contract','Clearer — it\'s the contracted bill of quantities'),
    ('Role merge', 'Detailing Head (separate role)',  'Merged into Team Lead','Same person. Detailing Head role deleted from system. Any user with detailing_head role → reassign to team_lead.'),
    ('Filter',     'Payments (Design/Services Head)', 'Advance + Final only','Design/Services Heads see only advance and final bill PRs — not the weekly payment cycle'),
    ('Remove',     'Vendors tab (at project setup only roles)','Accessible from project setup flow only','Team Lead, Jr Arch, Services Eng, Coordinator do not need persistent Vendors tab'),
    ('Move',       'GRN — was in Work for all roles','Money for Principal/PMCHead/DesignHead\nWork for Site/SeniorSite/Coordinator','GRN is money-chain for approvers, field action for site team'),
    ('Drawings',   'Principal had no Drawings tab',   'Drawing status in Project Summary only','Principals don\'t manage drawings day-to-day. Status visible as count + tap-through in Project Summary.'),
    ('Add',        'Drawings not in PMC Head Work',   'Added: Drawings + Register in PMC Work','PMC needs drawing visibility for NCRs, MOM references, site execution oversight'),
]

r=5
for change,old,new,note in CONS:
    fill = {'Merge':BG_WORK,'Rename':BG_HOME,'Filter':BG_MONEY,'Remove':BG_PEND,'Move':BG_MORE,'Role merge':BG_PEND,'Drawings':BG_HOME,'Add':BG_WORK}.get(change,BG_HIDE)
    cel(ws3,r,1,change,fill,Font(name='Arial',size=10,bold=True),C)
    cel(ws3,r,2,old,BG_HIDE,F_B,L)
    cel(ws3,r,3,new,BG_YEL,F_B,L)
    cel(ws3,r,4,note,BG_HIDE,Font(name='Arial',size=9,color='555555'),L)
    ws3.row_dimensions[r].height=46
    r+=1

# ═══════════════════════════════
# SHEET 4 — NEW FEATURES LOGGED
# ═══════════════════════════════
ws4 = wb.create_sheet('4 · New Features Logged')
ws4['A1']='nu PMC — New Features Arising from Nav Design'; ws4['A1'].font=F_T; ws4.merge_cells('A1:D1')
ws4['A2']='These are not nav changes — they are new product features to build separately.'; ws4['A2'].font=F_M; ws4.merge_cells('A2:D2')

for c,(v,w) in enumerate([('#',5),('Feature',30),('Description',60),('Priority',14)],1):
    hdr(ws4,4,c,v,w)

FEATURES=[
    (1,'Daily Report Submission by Site Manager',
     'Site manager submits end-of-day report by 6pm. Contains: tasks marked today + photos uploaded + issues raised + labour headcount + free text observations. PMC Head has overnight to review/correct. 6am: auto-publishes to principals + design head + services head + PMC head (in-app only, no WhatsApp). PMC correction replaces original; original preserved in audit log.',
     'High'),

    (2,'Project-Scoping Fix (security)',
     'Backend only scopes site_manager to assigned projects. Needs fix for: senior_site_manager, team_lead, jr_architect, services_engineer, coordinator, detailing, trainee. Fix in routes/projects.js (role check expansion) and routes/auth.js (store assigned projects in session for all 7 roles). Also needs project-membership check on all per-project endpoints for these roles (part of CAT 11 security audit findings).',
     'High'),

    (3,'Pending Tab (Blocked + Needs You)',
     'New tab for Principal, Design Principal, PMC Head. Blocked: things overdue in others\' queues (drawings with design head, RFIs, GRNs, MOMs, vendor clearance, project health red/amber). Needs You: items correctly routed to this person awaiting their action (PRs, engagements, claims, CNs, sign-off). Dual nudge: Day N (per-project SLA) → item appears in Pending + WhatsApp to owner. Day N+2 → WhatsApp to principal.',
     'High'),

    (4,'Per-Project SLA Configuration',
     'SLAs set during project setup. Default values pre-filled: GRN=2d, Drawings=3d, RFI=5d, Vendor Clearance=7d, MOM=3d, PR=2d. Editable later from Project Summary. Drives the Pending tab thresholds and dual-nudge timing. Requires new DB table: project_slas(project_id, item_type, sla_days).',
     'High'),

    (5,'DB-Driven Nav (ROLE_TABS migration)',
     'ROLE_TABS currently hardcoded JS array in app.js lines 13-36. Migrate to DB table role_nav(role, bucket, tab_key, sort_order, is_visible). Frontend fetches at login and builds nav dynamically. Enables IT Admin editor without code deploys.',
     'Medium'),

    (6,'IT Admin Nav Editor',
     'IT Admin can edit nav config via table UI: change bucket assignment, reorder within bucket, toggle visibility per tab per role. Changes saved as draft (not live). WhatsApp to Principal. Principal sees plain-English diff in Pending tab. Approve → goes live for all users of that role. Reject → draft discarded. Principal approval required before any change goes live.',
     'Medium'),

    (7,'Sort Toggle on All List Views',
     'Every tab with a list of items gets a 3-way sort toggle: Urgency / Age (oldest first) / Default (date added). Sort preference remembered per tab per role. Default varies: Principal=Age, PMC Head=Urgency, Design Head=Age, Site Manager=Default. Fixed ordering now, adaptive (learns from usage) later.',
     'Medium'),

    (8,'Project Summary Redesign (role-aware)',
     'Replace current bland Project Summary with role-specific buttons and counts. See separate Project Summary Audit Excel for full spec. Backend: extend GET /api/projects/:id to return role-aware summary block. Single call, faster on site wifi.',
     'Medium'),

    (9,'Needs You Pinned Section in Work',
     'For Design Head, Services Head, Team Lead, Jr Architect, Services Engineer, Coordinator, Senior Site Manager: pinned section at top of Work bucket showing pending items count by type. Hides when empty. Team Lead/Detailing Head: shows team\'s pending items. Others: own items only.',
     'Medium'),

    (10,'Today\'s Report Pinned in Work (Site Manager)',
     'Pinned section at top of Site Manager Work bucket. Shows submission status: Not submitted (due 6pm) / Submitted (time) / Published (6am). Tap to open report submission form. Pre-populates from tasks, photos, issues, labour already entered today.',
     'Medium'),
]

r=5
for num,name,desc,pri in FEATURES:
    fill={'High':BG_PEND,'Medium':BG_FLAG}.get(pri,BG_HIDE)
    cel(ws4,r,1,str(num),BG_SUB,F_S,C)
    cel(ws4,r,2,name,fill,Font(name='Arial',size=10,bold=True),LC)
    cel(ws4,r,3,desc,BG_HIDE,Font(name='Arial',size=9),L)
    cel(ws4,r,4,pri,fill,Font(name='Arial',size=10,bold=True,color='8B4513' if pri=='High' else NAVY),C)
    ws4.row_dimensions[r].height=90
    r+=1

# ═══════════════════════════════
# SHEET 5 — PROJECT SCOPING
# ═══════════════════════════════
ws5 = wb.create_sheet('5 · Project Scoping')
ws5['A1']='nu PMC — Project Scoping per Role'; ws5['A1'].font=F_T; ws5.merge_cells('A1:C1')
ws5['A2']='Project-scoped roles see only their assigned projects. Firm-wide roles see all.'; ws5['A2'].font=F_M; ws5.merge_cells('A2:C2')

for c,(v,w) in enumerate([('Role',24),('Scope',20),('Current bug?',20)],1):
    hdr(ws5,4,c,v,w)

SCOPING=[
    ('Principal',          'Firm-wide — all projects','No bug'),
    ('Design Principal',   'Firm-wide — all projects','No bug'),
    ('PMC Head',           'Firm-wide — all projects','No bug'),
    ('Design Head',        'Firm-wide — all projects','No bug'),
    ('Services Head',      'Firm-wide — all projects','No bug'),
    ('Finance Admin',      'Firm-wide — all projects','No bug'),
    ('Audit',              'Firm-wide — all projects (read-only)','No bug'),
    ('Site Manager',       'Project-scoped — assigned only','Fixed (enforced in code)'),
    ('Senior Site Manager','Project-scoped — assigned only','⚠ Bug: session has projects but backend query not scoped'),
    ('Team Lead',          'Project-scoped — assigned only','⚠ Bug: sees all projects'),
    ('Jr Architect',       'Project-scoped — assigned only','⚠ Bug: sees all projects'),
    ('Services Engineer',  'Project-scoped — assigned only','⚠ Bug: sees all projects'),
    ('Coordinator',        'Project-scoped — assigned only','⚠ Bug: sees all projects'),
    ('Detailing',          'Project-scoped — assigned only','⚠ Bug: sees all projects'),
    ('Trainee',            'Project-scoped — assigned only','⚠ Bug: sees all projects'),
]

r=5
for role,scope,bug in SCOPING:
    is_bug = '⚠' in bug
    cel(ws5,r,1,role,BG_SUB,F_S,LC)
    cel(ws5,r,2,scope,BG_HOME if 'Firm' in scope else BG_WORK,F_B,LC)
    cel(ws5,r,3,bug,BG_PEND if is_bug else BG_HIDE,Font(name='Arial',size=10,bold=is_bug,color='8B4513' if is_bug else '000000'),LC)
    ws5.row_dimensions[r].height=28
    r+=1

r+=2
ws5.cell(row=r,column=1,value='Fix required:').font=F_S
ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
r+=1
fix_text = ('routes/projects.js line 15: expand role check from site_manager only to all 7 scoped roles.\n'
            'routes/auth.js: store assigned projects in session for all 7 scoped roles (currently only site_manager + senior_site_manager).\n'
            'All per-project endpoints: add project-membership verification for scoped roles.')
c = ws5.cell(row=r, column=1, value=fix_text)
c.font = Font(name='Consolas', size=9); c.fill=BG_FLAG; c.alignment=L; c.border=BOX
ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
ws5.row_dimensions[r].height=72

import os
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
OUT = '/mnt/user-data/outputs/20260420 nu PMC NavRedesign Audit v1.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheets: {len(wb.sheetnames)}')
