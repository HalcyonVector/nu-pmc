from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── Styling ───
NAVY = '1D3D62'
BG_HEADER = PatternFill('solid', start_color=NAVY)
BG_SUBHEADER = PatternFill('solid', start_color='E8EEF4')
BG_NO = PatternFill('solid', start_color='F4F4F4')
BG_YES = PatternFill('solid', start_color='E8F4EE')
BG_YELLOW = PatternFill('solid', start_color='FFFBEF')

F_HEADER = Font(name='Arial', size=11, bold=True, color='FFFFFF')
F_SUB = Font(name='Arial', size=10, bold=True, color=NAVY)
F_BODY = Font(name='Arial', size=10)
F_BODY_BOLD = Font(name='Arial', size=10, bold=True)
F_MUTED = Font(name='Arial', size=9, color='888888', italic=True)

thin = Side(border_style='thin', color='CCCCCC')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
CENTER_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ════════════════════════════════════════════════
# SHEET 1 — Role × Button matrix
# ════════════════════════════════════════════════
ws = wb.active
ws.title = 'Role × Button'

ws['A1'] = 'nu PMC — Project Summary Audit'
ws['A1'].font = Font(name='Arial', size=14, bold=True, color=NAVY)
ws.merge_cells('A1:H1')

ws['A2'] = 'Mark changes directly in the cells. What each role sees when they tap "Project Summary" on the sidebar.'
ws['A2'].font = F_MUTED
ws.merge_cells('A2:H2')

headers = ['Role', 'Schedule', 'Issues / Project Issues / My Issues', 'CNs', 'Approvals',
           'Reports', 'Pending Submittals', "Today's Tasks / Payments Queue"]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col_idx, value=h)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = CENTER
    c.border = BOX

# Per-role rows — role titles only, no names
roles = [
    ('Principal',            'Yes', 'All open issues', 'Yes', 'Yes (all categories)',   '—', '—', '—'),
    ('Design Principal',     'Yes', 'All open issues', 'Yes', 'Yes (all categories)',   '—', '—', '—'),
    ('PMC Head',             '—',   'All open issues', 'Yes', 'Yes (all categories)',   'Yes (daily reports)', '—', '—'),
    ('Design Head',          '—',   'Design stream',   '—',   'Drawings + Budget (design only)', '—', '—', '—'),
    ('Services Head',        '—',   'Services stream', '—',   'Drawings + Budget (services only)', '—', '—', '—'),
    ('Finance Admin',        '—',   '—',               '—',   '—',                       '—', '—', 'Payments Queue'),
    ('Senior Site Manager',  'Yes', 'All open issues (project)', '—', 'Payments → GRN only', '—', '—', '—'),
    ('Site Manager',         'Yes', 'All open issues (project)', '—', 'Payments → GRN only', '—', '—', '—'),
    ('Team Lead',            '—',   'My issues only',  '—',   '—',                       '—', 'Yes', '—'),
    ('Detailing Head',       '—',   'My issues only',  '—',   '—',                       '—', 'Yes', '—'),
    ('Jr Architect',         '—',   'My issues only',  '—',   '—',                       '—', '—', '—'),
    ('Services Engineer',    '—',   'My issues only',  '—',   '—',                       '—', '—', '—'),
    ('Coordinator',          '—',   'My issues only',  '—',   '—',                       '—', '—', "Today's Tasks"),
    ('Detailing',            '—',   '—',               '—',   '—',                       '—', '—', '— (no project_detail tab)'),
    ('Trainee',              '—',   '—',               '—',   '—',                       '—', '—', '— (no project_detail tab)'),
    ('Audit',                'Yes', 'All open issues', 'Yes', 'Yes (all read-only)',     'Yes (read-only)', '—', '—'),
]

row = 5
for role_data in roles:
    for col_idx, val in enumerate(role_data, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.border = BOX
        c.alignment = LEFT
        if col_idx == 1:
            c.font = F_BODY_BOLD
            c.fill = BG_SUBHEADER
        else:
            c.font = F_BODY
            if val == '—':
                c.fill = BG_NO
                c.alignment = CENTER
            elif val.startswith('Yes'):
                c.fill = BG_YES
    ws.row_dimensions[row].height = 28
    row += 1

widths = [24, 10, 28, 8, 28, 22, 18, 26]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 42

row += 1
ws.cell(row=row, column=1, value='Audit notes').font = F_SUB
row += 1
notes = [
    'Legend: "Yes" = button visible on Project Summary for this role. "—" = button hidden.',
    'Schedule button: rich view (Today · Look Ahead · Date strip + Milestones). Principals + site managers.',
    'Issues button label varies — "Issues" for principals/PMC, "Project Issues" for site managers, "My Issues" for team_lead/jr_architect/services_engineer/coordinator.',
    'Approvals button shows 5 categories, filtered to what role can action. See sheet 2 for category details.',
    'Finance Admin sees ONE button only: Payments Queue. They do not need Issues/CNs on project summary (they have dedicated Payments tab).',
    'Detailing and Trainee roles have no Project Summary tab at all — removed from their sidebar.',
]
for n in notes:
    c = ws.cell(row=row, column=1, value='• ' + n)
    c.font = F_BODY
    c.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 24
    row += 1

# ════════════════════════════════════════════════
# SHEET 2 — Approvals categories
# ════════════════════════════════════════════════
ws2 = wb.create_sheet('Approvals categories')
ws2['A1'] = 'Approvals button — 5 categories'
ws2['A1'].font = Font(name='Arial', size=14, bold=True, color=NAVY)
ws2.merge_cells('A1:C1')
ws2['A2'] = 'What appears inside each category, and who sees it.'
ws2['A2'].font = F_MUTED
ws2.merge_cells('A2:C2')

headers2 = ['Category', "What's inside (auto-filtered per role)", 'Who sees this category']
for col_idx, h in enumerate(headers2, 1):
    c = ws2.cell(row=4, column=col_idx, value=h)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = CENTER
    c.border = BOX

cats = [
    ('Drawings',
     '• Drawing version approvals (design stream → design_head, services → services_head)\n'
     '• Submittal reviews\n'
     '• Change Notice approvals (peer below ₹1L, principal above)',
     'Principal, Design Principal, PMC Head, Design Head (design only), Services Head (services only), Audit'),
    ('Payments',
     '• GRN approvals (senior site below 5% budget, PMC any value, principal any)\n'
     '• Payment Request PMC-review\n'
     '• Payment Request naveen-review\n'
     '• Vendor engagement approvals\n'
     '• Client claim approvals',
     'Principal, Design Principal (all), PMC Head (GRN + PR), Senior Site / Site Manager (GRN only), Audit'),
    ('Budget',
     '• Custom cost-head approvals (principal + design_head + services_head)\n'
     '• Budget flag sign-offs (PMC + stream heads)',
     'Principal, Design Principal, PMC Head, Design Head (design only), Services Head (services only), Audit'),
    ('MOMs',
     '• Draft MOMs awaiting PMC Head approval before issue to client\n'
     '  (internal checkpoint — separate step then "issue to client")',
     'PMC Head (approves), Principal, Design Principal (visibility), Audit'),
    ('Other',
     "• Schedule change approvals routed via approvals queue\n"
     '• Weekly report approvals (rare)\n'
     "• Anything else that doesn't fit above 4",
     'Principal, Design Principal, PMC Head, Audit'),
]

row = 5
for cat, inside, who in cats:
    c = ws2.cell(row=row, column=1, value=cat)
    c.font = F_BODY_BOLD
    c.alignment = CENTER_LEFT
    c.fill = BG_SUBHEADER
    c.border = BOX
    ws2.cell(row=row, column=2, value=inside).font = F_BODY
    ws2.cell(row=row, column=2).alignment = LEFT
    ws2.cell(row=row, column=2).border = BOX
    ws2.cell(row=row, column=3, value=who).font = F_BODY
    ws2.cell(row=row, column=3).alignment = LEFT
    ws2.cell(row=row, column=3).border = BOX
    ws2.row_dimensions[row].height = 90
    row += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 64
ws2.column_dimensions['C'].width = 56

# ════════════════════════════════════════════════
# SHEET 3 — Change summary
# ════════════════════════════════════════════════
ws3 = wb.create_sheet('Change summary')
ws3['A1'] = 'What changes in this revision'
ws3['A1'].font = Font(name='Arial', size=14, bold=True, color=NAVY)
ws3.merge_cells('A1:C1')
ws3['A2'] = 'How the Project Summary screen shifts from current (bland) to revised (role-aware)'
ws3['A2'].font = F_MUTED
ws3.merge_cells('A2:C2')

headers3 = ['Area', 'Current (what you saw in preview)', 'Revised (what this spec delivers)']
for col_idx, h in enumerate(headers3, 1):
    c = ws3.cell(row=4, column=col_idx, value=h)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = CENTER
    c.border = BOX

changes = [
    ('Team roster',
     '11-person list shown to everyone',
     'Dropped for all roles'),
    ('Issues count',
     'Single number, same for all',
     'Role-filtered — stream-scoped for design/services heads, all for PMC/principals, assigned-only for team_lead/jr_architect'),
    ('CNs count',
     'Single number, same for all',
     "Only visible to Principal + Design Principal + PMC Head + Audit (others don't approve CNs)"),
    ('Team count stat',
     'Always shown ("11 Team")',
     'Dropped — redundant when team roster is gone'),
    ('Approvals button',
     'Did not exist',
     'New button on Project Summary. Tap → 5-tab strip (Drawings / Payments / Budget / MOMs / Other). Each tab auto-filtered by role.'),
    ('Schedule button',
     'Did not exist',
     'New button for Principals + Site Managers. Tap → rich view (Today · Look Ahead · date strip + milestones)'),
    ('Finance Admin',
     'Sees generic Issues/CNs/Team',
     'Sees one button: Payments Queue (route to existing payments-fin tab)'),
    ('Site Manager',
     'Generic same-for-all view',
     'Schedule + Project Issues + Approvals (GRN only)'),
    ('Data fetch',
     'Single /api/projects/:id call returns static project data',
     'Same endpoint extended — returns role-aware summary block with button list, counts, category breakdowns'),
    ('Backend change',
     'None needed',
     'Extend /api/projects/:id handler. New summary builder in routes/projects.js. ~2.5 hrs work.'),
    ('Frontend change',
     'None needed',
     'Rewrite APP.renderProjectDetail. New sub-view for Approvals tab-strip. ~1.5 hrs.'),
    ('Static project header',
     'Name, client, location, contract value',
     'SAME — unchanged. Stays at top of Project Summary.'),
]

row = 5
for area, cur, rev in changes:
    c = ws3.cell(row=row, column=1, value=area)
    c.font = F_BODY_BOLD
    c.alignment = CENTER_LEFT
    c.fill = BG_SUBHEADER
    c.border = BOX
    ws3.cell(row=row, column=2, value=cur).font = F_BODY
    ws3.cell(row=row, column=2).alignment = LEFT
    ws3.cell(row=row, column=2).border = BOX
    ws3.cell(row=row, column=3, value=rev).font = F_BODY
    ws3.cell(row=row, column=3).alignment = LEFT
    ws3.cell(row=row, column=3).border = BOX
    ws3.row_dimensions[row].height = 50
    row += 1

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 70

# ════════════════════════════════════════════════
# SHEET 4 — Your sign-off
# ════════════════════════════════════════════════
ws4 = wb.create_sheet('Your sign-off')
ws4['A1'] = 'Audit sign-off'
ws4['A1'].font = Font(name='Arial', size=14, bold=True, color=NAVY)
ws4.merge_cells('A1:D1')
ws4['A2'] = 'Mark each row Approved / Change / Reject and add comments as needed. Send back to me.'
ws4['A2'].font = F_MUTED
ws4.merge_cells('A2:D2')

headers4 = ['#', 'Item', 'Decision (Approved / Change / Reject)', 'Comment / what to change']
for col_idx, h in enumerate(headers4, 1):
    c = ws4.cell(row=4, column=col_idx, value=h)
    c.font = F_HEADER
    c.fill = BG_HEADER
    c.alignment = CENTER
    c.border = BOX

items = [
    'Role × Button matrix as shown on Sheet 1',
    'Approvals 5 categories (Drawings / Payments / Budget / MOMs / Other)',
    'Team roster dropped for all roles',
    'Schedule button for Principals + Site Managers',
    'Finance Admin sees only Payments Queue',
    'Site Manager sees Schedule + Project Issues + Approvals',
    'Team Lead / Detailing Head see My Issues + Pending Submittals',
    "Coordinator sees Today's Tasks + My Issues",
    'Jr Architect / Services Engineer see only My Issues',
    'Detailing / Trainee have no Project Summary tab at all',
    'Audit role sees everything read-only',
    'Backend extends /api/projects/:id (single call)',
    'Estimated build time: ~5 hours',
]

row = 5
for idx, item in enumerate(items, 1):
    c = ws4.cell(row=row, column=1, value=idx)
    c.font = F_BODY_BOLD
    c.alignment = CENTER
    c.fill = BG_SUBHEADER
    c.border = BOX
    ws4.cell(row=row, column=2, value=item).font = F_BODY
    ws4.cell(row=row, column=2).alignment = LEFT
    ws4.cell(row=row, column=2).border = BOX
    for col in (3, 4):
        c = ws4.cell(row=row, column=col, value='')
        c.font = F_BODY
        c.alignment = LEFT if col == 4 else CENTER
        c.border = BOX
        c.fill = BG_YELLOW
    ws4.row_dimensions[row].height = 30
    row += 1

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 60
ws4.column_dimensions['C'].width = 32
ws4.column_dimensions['D'].width = 50

import os
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
OUT = '/mnt/user-data/outputs/20260420 nu ProjectSummary Audit v2.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')
