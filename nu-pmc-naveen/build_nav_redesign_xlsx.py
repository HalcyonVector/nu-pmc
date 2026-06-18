from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ─── Styles ───
NAVY     = '1D3D62'
BG_NAV   = PatternFill('solid', start_color=NAVY)
BG_HOME  = PatternFill('solid', start_color='D6E4F5')
BG_WORK  = PatternFill('solid', start_color='FFF0D6')
BG_MONEY = PatternFill('solid', start_color='D6EFE0')
BG_MORE  = PatternFill('solid', start_color='EDE8DF')
BG_PEND  = PatternFill('solid', start_color='F5D6D6')
BG_STRIP = PatternFill('solid', start_color='F0EDE8')
BG_HIDE  = PatternFill('solid', start_color='F4F4F4')
BG_SUB   = PatternFill('solid', start_color='EEF3F9')
BG_YELL  = PatternFill('solid', start_color='FFFBEF')
BG_FLAG  = PatternFill('solid', start_color='FFF3CD')

BFILL = {'Home':BG_HOME,'Work':BG_WORK,'Money':BG_MONEY,'More':BG_MORE,
         'Pending':BG_PEND,'2-tab strip':BG_STRIP,'—':BG_HIDE}

F_TITLE = Font(name='Arial', size=14, bold=True, color=NAVY)
F_HEAD  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_SUB   = Font(name='Arial', size=10, bold=True, color=NAVY)
F_BODY  = Font(name='Arial', size=10)
F_SMALL = Font(name='Arial', size=9, color='444444')
F_MUTED = Font(name='Arial', size=9, color='888888', italic=True)
F_FLAG  = Font(name='Arial', size=9, bold=True, color='8B4513')

thin = Side(border_style='thin', color='CCCCCC')
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)
C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
L  = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
LC = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def h(ws, row, col, val, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font=F_HEAD; c.fill=BG_NAV; c.alignment=C; c.border=BOX
    if w: ws.column_dimensions[get_column_letter(col)].width = w
    return c

def cl(ws, row, col, val, fill=None, font=None, align=None):
    if fill is None: fill = BG_HIDE
    c = ws.cell(row=row, column=col, value=val)
    c.font=font or F_BODY; c.fill=fill; c.alignment=align or L; c.border=BOX
    return c

# ═══════════════════════════════════════════════
# SHEET 1 — MASTER NAV (bottom bar per role)
# ═══════════════════════════════════════════════
ws = wb.create_sheet('1 · Master Nav')
ws['A1'] = 'nu PMC — Navigation Redesign · Bottom Bar per Role'
ws['A1'].font = F_TITLE; ws.merge_cells('A1:G1')
ws['A2'] = 'Locked decisions. Yellow row = your audit notes. No build until signed off.'
ws['A2'].font = F_MUTED; ws.merge_cells('A2:G2')

for col, (v, w) in enumerate([('Role',24),('Button 1',14),('Button 2',14),('Button 3',14),('Button 4',14),('Bars visible',12),('Notes / Special behaviour',42)], 1):
    h(ws, 4, col, v, w)
ws.row_dimensions[4].height = 30

ROLES = [
    ('Principal',           'Home','Money','Pending','More',    4, 'Pending = Blocked + Needs You. Drawings via Project Summary only.'),
    ('Design Principal',    'Home','Money','Pending','More',    4, 'Identical to Principal.'),
    ('PMC Head',            'Home','Work', 'Money',  'Pending', 4, 'Needs You pinned top of Work. Pending = Blocked (site team overdue).'),
    ('Design Head',         'Home','Work', 'Money',  'More',    4, 'Needs You pinned top of Work. Payments = advance+final only.'),
    ('Services Head',       'Home','Work', 'Money',  'More',    4, 'Same as Design Head, services stream filter.'),
    ('Site Manager',        'Home','Work', 'Money',  '—',       3, 'Today\'s Report pinned top of Work (submit by 6pm).'),
    ('Senior Site Manager', 'Home','Work', 'Money',  '—',       3, 'Needs You (GRN approvals) + Today\'s Report both pinned in Work.'),
    ('Finance Admin',       'Home','Money','—',      '—',       2, 'Udupa only. Entire working life is in Money.'),
    ('Team Lead',           'Home','Work', '—',      '—',       2, 'Team Needs You pinned. Sees team\'s pending items. (Detailing Head merged here)'),
    ('Jr Architect',        'Home','Work', '—',      '—',       2, 'My Needs You pinned. Own items only.'),
    ('Services Engineer',   'Home','Work', '—',      '—',       2, 'My Needs You pinned. Own items only.'),
    ('Coordinator',         'Home','Work', '—',      '—',       2, 'Needs You pinned.'),
    ('Trainee',         '2-tab strip','—','—',       '—',       0, 'No bottom bar. Top strip only: Drawings · Schedule.'),
    ('Detailing',       '2-tab strip','—','—',       '—',       0, 'No bottom bar. Top strip only: Drawings · Submittals.'),
    ('Audit',               'Home','Money','Pending','More',    4, 'Mirrors Principal. All action buttons hidden — read only.'),
]

row = 5
for role, b1, b2, b3, b4, bars, note in ROLES:
    cl(ws, row, 1, role, BG_SUB, F_SUB, LC)
    for col, val in [(2,b1),(3,b2),(4,b3),(5,b4)]:
        cl(ws, row, col, val, BFILL.get(val,BG_HIDE), F_BODY, C)
    cl(ws, row, 6, bars if bars > 0 else 'strip', BG_HIDE, F_BODY, C)
    cl(ws, row, 7, note, None, F_SMALL, L)
    ws.row_dimensions[row].height = 28
    row += 1

row += 1
ws.cell(row=row, column=1, value='Your audit notes:').font = F_SUB
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
for col in range(1, 8):
    c = ws.cell(row=row, column=col); c.fill=BG_YELL; c.border=BOX
ws.row_dimensions[row].height = 60

# ═══════════════════════════════════════════════
# SHEET 2 — BUCKET CONTENTS per role
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('2 · Bucket Contents')
ws2['A1'] = 'nu PMC — What\'s Inside Each Bucket · Per Role'
ws2['A1'].font = F_TITLE; ws2.merge_cells('A1:E1')
ws2['A2'] = 'Tabs listed frequency order: most used → least used. Sort toggle on every list: Urgency / Age / Default.'
ws2['A2'].font = F_MUTED; ws2.merge_cells('A2:E2')

for col, (v, w) in enumerate([('Role',22),('🏠 Home',24),('🔨 Work',36),('₹ Money',28),('More / Pending',34)], 1):
    h(ws2, 4, col, v, w)
ws2.row_dimensions[4].height = 30

CONTENTS = [
    ('Principal',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary\n(Drawings visible inside Project Summary — no separate tab)',
     '—\n(principals review + approve, not execute)',
     'Payments\nVendors (Master+Clearance merged)\nBudget (Summary+Tree toggle inside)\nVendor Allocation (renamed BOQ Map)\nClient Contract (renamed Client BOQ)',
     'MORE:\nDelegations\nCN Register\nHealth Dashboard\nUsers\n\nPENDING:\n🔴 Blocked — drawings overdue/Design Head, RFIs overdue, GRNs overdue/PMC, MOMs not approved, vendor clearance overdue, project health red/amber\n📋 Needs You — PR naveen-review, engagement approvals, client claims, CNs above ₹1L, schedule changes, weekly sign-off'),

    ('Design Principal',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary',
     '—',
     'Payments\nVendors\nBudget\nVendor Allocation\nClient Contract',
     'MORE: same as Principal\nPENDING: same as Principal'),

    ('PMC Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  Reports to approve · N pending\n  MOMs to approve · N pending\n  PRs to PMC-review · N pending\n  GRNs to sign off · N pending\n─────────────\nReports\nIssues\nMeetings\nDrawings\nRegister\nMaterials\nLabour',
     'GRNs\nPayments\nVendors\nVendor Master',
     'PENDING (4th button):\n🔴 Blocked:\n  Daily reports not submitted by site\n  Issues assigned to site — no update\n  Labour headcount not entered\n  Photos not uploaded\n  (all measured against per-project SLAs)'),

    ('Design Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  Drawing versions pending review\n  Submittals to review\n  Material approvals pending\n─────────────\nDrawings\nIssues\nSubmittals\nRegister\nPhoto Review',
     'Materials\nBudget\nPayments (advance + final only)',
     'MORE:\nWeekly Sign-off\nDelegations'),

    ('Services Head',
     'Dashboard\nMonthly Overview\nProject Summary',
     '⚡ NEEDS YOU (pinned)\n[services stream filter]\n─────────────\nDrawings\nIssues\nSubmittals\nRegister\nPhoto Review',
     'Materials\nBudget\nPayments (advance + final only)',
     'MORE:\nWeekly Sign-off\nDelegations'),

    ('Site Manager',
     'Dashboard\nProject Summary\n(no Monthly — site-scoped only)',
     '📋 TODAY\'S REPORT (pinned):\n  Status: Not submitted / Submitted / Published\n  Due: 6pm daily\n─────────────\nTasks (assigned to me, today)\nPhotos\nField Issues\nIssues\nLabour\nDrawings (read-only, latest approved)\nRegister (read-only)',
     'GRNs (raise when materials arrive)\nPayments (raise PR with invoice photo)\nMaterials (delivery status)',
     '—\n(no More or Pending tab)'),

    ('Senior Site Manager',
     'Dashboard\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  GRN approvals below 5% budget\n📋 TODAY\'S REPORT (pinned)\n─────────────\nTasks\nPhotos\nField Issues\nIssues\nLabour\nDrawings (read-only)\nRegister (read-only)',
     'GRNs\nPayments\nMaterials',
     '—'),

    ('Finance Admin',
     'Dashboard\nMonthly Overview\nProject Summary',
     '—\n(no Work — finance is not field execution)',
     'Payments (Finance) — Saturday workflow\nVendors (Master+Clearance merged)\nClient Receipts\nPetty Cash\nInvoices (PI)\nGST Statement\nClient Contract\nClients',
     '—\n(no More or Pending)'),

    ('Team Lead',
     'Dashboard\nProject Summary',
     '⚡ TEAM NEEDS YOU (pinned):\n  Team\'s pending RFIs\n  Submittals returned for resubmit\n  Drawings pending upload by team\n─────────────\nDrawings\nIssues\nSubmittals\nRegister',
     '—',
     '—\n(Detailing Head merged into this role)'),

    ('Jr Architect',
     'Dashboard\nProject Summary',
     '⚡ MY NEEDS YOU (pinned):\n  RFIs assigned to me\n  Submittals returned to me\n─────────────\nDrawings\nIssues\nSubmittals',
     '—',
     '—'),

    ('Services Engineer',
     'Dashboard\nProject Summary',
     '⚡ MY NEEDS YOU (pinned)\n─────────────\nDrawings\nIssues\nSubmittals\nPhoto Review',
     '—',
     '—'),

    ('Coordinator',
     'Dashboard\nProject Summary',
     '⚡ NEEDS YOU (pinned):\n  MOM action items assigned to me\n  Issues assigned to me\n─────────────\nMeetings\nTasks\nIssues\nDrawings\nRegister\nPhotos\nGRNs',
     '—',
     '—'),

    ('Trainee',
     '—',
     'NO BOTTOM BAR — 2-tab top strip:\n• Drawings (read-only, approved versions)\n• Schedule (read-only)',
     '—',
     '—'),

    ('Detailing',
     '—',
     'NO BOTTOM BAR — 2-tab top strip:\n• Drawings (upload detail drawings)\n• Submittals (upload)',
     '—',
     '—'),

    ('Audit',
     'Dashboard\nMonthly Overview\nProjects\nProject Summary',
     'All Work tabs — READ ONLY\n(no action buttons, no submit, no approve)',
     'All Money tabs — READ ONLY',
     'MORE + PENDING:\nAll tabs visible\nAll action buttons hidden\nAudit trail visible'),
]

row = 5
for items in CONTENTS:
    role = items[0]
    cols_data = items[1:]
    fills = [BG_SUB, BG_HOME, BG_WORK, BG_MONEY, BG_MORE]
    fonts = [F_SUB, F_SMALL, F_SMALL, F_SMALL, F_SMALL]
    aligns = [LC, L, L, L, L]
    for col_idx, (val, fill, font, align) in enumerate(zip(cols_data, fills, fonts, aligns), 2):
        cl(ws2, row, col_idx, val, fill, font, align)
    cl(ws2, row, 1, role, BG_SUB, F_SUB, LC)
    lines = max(v.count('\n') for v in cols_data) + 1
    ws2.row_dimensions[row].height = max(80, lines * 14)
    row += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 26
ws2.column_dimensions['C'].width = 38
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 38

# ═══════════════════════════════════════════════
# SHEET 3 — CONSOLIDATIONS & RENAMES
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('3 · Consolidations')
ws3['A1'] = 'nu PMC — Tab Consolidations & Renames'
ws3['A1'].font = F_TITLE; ws3.merge_cells('A1:D1')
ws3['A2'] = 'Decisions locked. These reduce tab count and clarify naming before build.'
ws3['A2'].font = F_MUTED; ws3.merge_cells('A2:D2')

for col, (v, w) in enumerate([('Item',28),('Before',26),('After',26),('Note',38)], 1):
    h(ws3, 4, col, v, w)
ws3.row_dimensions[4].height = 28

CONSOL = [
    ('MERGE', 'Vendor Master + Vendor Clearance (2 separate tabs)', 'Vendors (1 tab, clearance is a section inside)', 'Clearance was a filtered sub-view of Master. No real distinction.'),
    ('MERGE', 'Budget + Budget Tree (2 separate tabs)', 'Budget (1 tab, Summary / Tree toggle inside)', 'Same data, two presentations. Toggle is simpler than 2 tabs.'),
    ('RENAME', 'BOQ Map', 'Vendor Allocation', '"BOQ Map" sounded like a map of the BOQ. Actual function is allocating BOQ items to vendors.'),
    ('RENAME', 'Client BOQ', 'Client Contract', 'More accurately describes the document — it is the client\'s contracted scope.'),
    ('FILTER', 'Payments (full tab for Design/Services Heads)', 'Payments — advance + final bill only', 'Heads see only the payment types relevant to their sign-off. Weekly cycle payments are PMC/principal territory.'),
    ('SCOPE', 'GRN — shown in same place for all roles', 'GRN in Money for Principal/PMCHead/DesignHead (approval). GRN in Work for SiteManager/SeniorSite/Coordinator (raise).', 'Same tab, different bucket placement based on what the role does with it.'),
    ('DELETE', 'Detailing Head (separate role)', 'Merged into Team Lead', 'Same person. One role, same nav structure as Team Lead.'),
    ('NEW', 'No daily report submission by site manager', 'Site manager submits report by 6pm. PMC corrects overnight. Auto-publishes 6am in-app to principals+heads.', 'New feature — not built yet. Site manager owns the daily record.'),
    ('BUG FIX', 'Project scoping — only site_manager filtered in backend', 'All 7 field roles scoped: senior_site, team_lead, jr_architect, services_eng, coordinator, detailing, trainee', 'Security gap — field roles could see all projects. Fix in routes/projects.js + routes/auth.js.'),
    ('NEW', 'ROLE_TABS hardcoded in app.js', 'Migrate to DB table role_nav(role, bucket, tab_key, sort_order). Admin screen for nav changes without redeploy.', 'Not built yet. Needed for nav redesign to be maintainable.'),
    ('NEW', 'No escalation SLAs', 'Per-project SLA table: GRN=2d, Drawings=3d, RFI=5d, Clearance=7d, MOM=3d, PR=2d. Set at project setup.', 'Drives Pending tab (Blocked) + dual nudge WhatsApp system.'),
    ('NEW', 'Sort order fixed (by date)', 'Sort toggle on every list: Urgency / Age / Default. Default varies per role.', 'Fixed ordering now. Adaptive (learns usage) deferred to later.'),
]

MERGE_FILL = {'MERGE':PatternFill('solid',start_color='D6EFE0'),'RENAME':PatternFill('solid',start_color='D6E4F5'),'FILTER':PatternFill('solid',start_color='FFF0D6'),'SCOPE':PatternFill('solid',start_color='EDE8DF'),'DELETE':PatternFill('solid',start_color='F5D6D6'),'NEW':PatternFill('solid',start_color='FFF3CD'),'BUG FIX':PatternFill('solid',start_color='FFE4E1')}

row = 5
for action, before, after, note in CONSOL:
    cl(ws3, row, 1, action, MERGE_FILL.get(action, BG_HIDE), F_SUB, C)
    cl(ws3, row, 2, before, BG_HIDE, F_SMALL, L)
    cl(ws3, row, 3, after, BG_HIDE, F_SMALL, L)
    cl(ws3, row, 4, note, BG_HIDE, F_SMALL, L)
    lines = max(before.count('\n'), after.count('\n'), note.count('\n')) + 1
    ws3.row_dimensions[row].height = max(42, lines * 15)
    row += 1

# ═══════════════════════════════════════════════
# SHEET 4 — PINNED SECTIONS (Needs You / Today's Report)
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('4 · Pinned Sections')
ws4['A1'] = 'nu PMC — Pinned Sections Inside Work Bucket'
ws4['A1'].font = F_TITLE; ws4.merge_cells('A1:D1')
ws4['A2'] = 'These appear at the TOP of the Work bucket. Hide when empty (count = 0). Tap row to jump to that filtered list.'
ws4['A2'].font = F_MUTED; ws4.merge_cells('A2:D2')

for col, (v, w) in enumerate([('Role',22),('Pinned Section(s)',24),('What it contains',38),('Behaviour',34)], 1):
    h(ws4, 4, col, v, w)
ws4.row_dimensions[4].height = 28

PINNED = [
    ('Principal / Design Principal', '(none — they have a separate Pending tab)', '—', 'Principals use the Pending bottom-bar button for triage, not a pinned section inside Work'),
    ('PMC Head', '⚡ Needs You', 'Reports to approve (count)\nMOMs to approve (count)\nPRs to PMC-review (count)\nGRNs to sign off (count)', 'Tap any row → filtered list of that type\nClears when count = 0\nCount shown as badge on Work button too'),
    ('Design Head', '⚡ Needs You', 'Drawing versions pending review\nSubmittals to review\nMaterial approvals pending', 'Design stream only\nHides when all clear'),
    ('Services Head', '⚡ Needs You', 'Drawing versions pending review\nSubmittals to review\nMaterial approvals pending', 'Services stream only\nHides when all clear'),
    ('Site Manager', '📋 Today\'s Report', 'Status: Not submitted / Submitted 5:43pm / Published\nTasks marked today: N of M\nPhotos uploaded: N\nIssues raised: N\nLabour headcount: N workers\nObservations: [free text]\n[Submit Report] button', 'Pinned always — site manager must see it every time\nGreen when submitted\nAmber warning if approaching 6pm deadline\nAfter 6am auto-publish → shows Published'),
    ('Senior Site Manager', '⚡ Needs You\n+\n📋 Today\'s Report', 'Needs You: GRN approvals below 5% budget\nToday\'s Report: same as site manager', 'Two pinned sections stacked\nNeeds You on top (action first)\nToday\'s Report below it'),
    ('Team Lead', '⚡ Team Needs You', 'Team\'s pending RFIs (not just own)\nSubmittals returned for resubmit\nDrawings pending upload by team members', 'Shows team items — not just own\nHides when team is all clear'),
    ('Jr Architect', '⚡ My Needs You', 'RFIs assigned to me awaiting my response\nSubmittals returned to me for resubmit', 'Own items only\nHides when clear'),
    ('Services Engineer', '⚡ My Needs You', 'Same as Jr Architect\n(services stream)', 'Own items only'),
    ('Coordinator', '⚡ Needs You', 'MOM action items assigned to coordinator\nIssues assigned to coordinator', 'Hides when clear'),
]

row = 5
for items in PINNED:
    role, pinned, contains, behaviour = items
    cl(ws4, row, 1, role, BG_SUB, F_SUB, LC)
    cl(ws4, row, 2, pinned, BG_WORK, F_BODY, L)
    cl(ws4, row, 3, contains, BG_HIDE, F_SMALL, L)
    cl(ws4, row, 4, behaviour, BG_HIDE, F_SMALL, L)
    lines = max(pinned.count('\n'), contains.count('\n'), behaviour.count('\n')) + 2
    ws4.row_dimensions[row].height = max(42, lines * 14)
    row += 1

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 24
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 36

# ═══════════════════════════════════════════════
# SHEET 5 — PENDING TAB DETAIL (Blocked + Needs You)
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('5 · Pending Tab')
ws5['A1'] = 'nu PMC — Pending Tab Detail (Principal + Design Principal + Audit)'
ws5['A1'].font = F_TITLE; ws5.merge_cells('A1:D1')
ws5['A2'] = 'Pending tab = 4th bottom-bar button. Contains two sub-views: Blocked + Needs You. PMC Head gets Blocked only as Pending tab (Needs You is pinned in Work).'
ws5['A2'].font = F_MUTED; ws5.merge_cells('A2:D2')

for col, (v, w) in enumerate([('Sub-view',18),('Item type',30),('Trigger condition',34),('Action available',28)], 1):
    h(ws5, 4, col, v, w)
ws5.row_dimensions[4].height = 28

PENDING = [
    ('🔴 BLOCKED', 'Drawing version stuck with Design Head', 'drawing_versions.status=pending AND age > project SLA (default 3d)', 'Nudge Rajani / Override / Approve yourself'),
    ('🔴 BLOCKED', 'RFI no response from Design Head', 'issues.type=rfi AND status=open AND age > SLA (5d)', 'Nudge / Reassign / Close'),
    ('🔴 BLOCKED', 'GRN pending PMC approval', 'grns.status=pending AND age > SLA (2d)', 'Approve / Escalate'),
    ('🔴 BLOCKED', 'MOM draft not approved by PMC Head', 'meetings.status=draft AND age > SLA (3d)', 'Nudge PMC / Approve yourself'),
    ('🔴 BLOCKED', 'Vendor clearance overdue with Finance', 'vendors.clearance_status=pending AND age > SLA (7d)', 'Nudge Udupa / Clear yourself'),
    ('🔴 BLOCKED', 'Submittal review overdue', 'submittals.status=pending AND age > SLA (5d)', 'Nudge head / Approve'),
    ('🔴 BLOCKED', 'Project health red or amber', 'weekly_health.score <= amber threshold', 'Tap → Health detail'),
    ('📋 NEEDS YOU', 'PR awaiting naveen-review', 'payment_requests.status=pending_naveen', 'Approve / Reject / Reduce amount'),
    ('📋 NEEDS YOU', 'Vendor engagement to approve', 'vendor_engagements.approval_status=pending', 'Approve / Reject with reason'),
    ('📋 NEEDS YOU', 'Client claim to approve', 'client_claims.status=pending_principal', 'Approve / Reject'),
    ('📋 NEEDS YOU', 'CN above ₹1L to approve', 'change_notices.status=pending AND amount > 100000', 'Approve / Reject'),
    ('📋 NEEDS YOU', 'Schedule change pending', 'approvals queue: schedule_change type', 'Approve / Reject'),
    ('📋 NEEDS YOU', 'Weekly sign-off pending', 'weekly_signoff.status=pending', 'Sign off'),
]

row = 5
prev_sub = None
for sub, item, trigger, action in PENDING:
    fill = BG_PEND if '🔴' in sub else BG_HOME
    if sub != prev_sub:
        cl(ws5, row, 1, sub, fill, F_SUB, LC)
    else:
        cl(ws5, row, 1, '', fill, F_BODY, C)
    cl(ws5, row, 2, item, BG_HIDE, F_SMALL, L)
    cl(ws5, row, 3, trigger, BG_HIDE, F_SMALL, L)
    cl(ws5, row, 4, action, BG_HIDE, F_SMALL, L)
    ws5.row_dimensions[row].height = 32
    prev_sub = sub
    row += 1

ws5.column_dimensions['A'].width = 18
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 38
ws5.column_dimensions['D'].width = 30

# ═══════════════════════════════════════════════
# SHEET 6 — SLA DEFAULTS
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet('6 · SLA Defaults')
ws6['A1'] = 'nu PMC — Escalation SLA Defaults (per-project, set at project setup)'
ws6['A1'].font = F_TITLE; ws6.merge_cells('A1:D1')
ws6['A2'] = 'These are default values. Each project can override at setup. After SLA days → item appears in Pending (Blocked) + WhatsApp nudge to owner. After SLA+2 days → WhatsApp to principal too.'
ws6['A2'].font = F_MUTED; ws6.merge_cells('A2:D2')

for col, (v, w) in enumerate([('Item type',28),('Default SLA (days)',20),('Who gets nudged on Day N',30),('Who gets nudged on Day N+2',30)], 1):
    h(ws6, 4, col, v, w)
ws6.row_dimensions[4].height = 28

SLAS = [
    ('GRN approval',           2, 'PMC Head',          'Principal'),
    ('Drawing version review',  3, 'Design/Services Head', 'Principal'),
    ('RFI response',            5, 'Design Head',       'Principal'),
    ('Vendor clearance',        7, 'Finance Admin',     'Principal'),
    ('MOM approval',            3, 'PMC Head',          'Principal'),
    ('PR PMC-review',           2, 'PMC Head',          'Principal'),
    ('Submittal review',        5, 'Design/Services Head', 'Principal'),
    ('Daily report submission', 1, 'Site Manager',      'PMC Head'),
    ('Labour headcount entry',  1, 'Site Manager',      'PMC Head'),
]

row = 5
for item, days, nudge1, nudge2 in SLAS:
    cl(ws6, row, 1, item, BG_SUB, F_BODY, LC)
    cl(ws6, row, 2, f'{days} working days', BG_PEND, F_BODY, C)
    cl(ws6, row, 3, nudge1, BG_HIDE, F_SMALL, LC)
    cl(ws6, row, 4, nudge2, BG_HIDE, F_SMALL, LC)
    ws6.row_dimensions[row].height = 28
    row += 1

row += 1
ws6.cell(row=row, column=1, value='⚠ SLAs are per-project. Fast projects (WeSchool ELCITA deadline) can tighten to 1d/2d. Slow projects can relax. Set at project initialisation. Editable from Project Summary later.').font = F_FLAG
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws6.row_dimensions[row].height = 36
ws6.cell(row=row, column=1).fill = BG_FLAG

ws6.column_dimensions['A'].width = 28
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 32
ws6.column_dimensions['D'].width = 32

# ═══════════════════════════════════════════════
# SHEET 7 — YOUR SIGN-OFF
# ═══════════════════════════════════════════════
ws7 = wb.create_sheet('7 · Sign-off')
ws7['A1'] = 'nu PMC Nav Redesign — Your Sign-off'
ws7['A1'].font = F_TITLE; ws7.merge_cells('A1:D1')
ws7['A2'] = 'Mark each item. Send back. No build until all items are Approved or Deferred.'
ws7['A2'].font = F_MUTED; ws7.merge_cells('A2:D2')

for col, (v, w) in enumerate([('#',5),('Decision item',52),('Approved / Change / Defer',26),('Your notes',40)], 1):
    h(ws7, 4, col, v, w)
ws7.row_dimensions[4].height = 28

ITEMS = [
    'Master nav table — 15 roles, 4-button max, as shown in Sheet 1',
    'Bucket contents per role — tabs + frequency order, as shown in Sheet 2',
    'Consolidations: Vendor Master+Clearance → one tab',
    'Consolidations: Budget+Budget Tree → one tab with toggle',
    'Renames: BOQ Map → Vendor Allocation, Client BOQ → Client Contract',
    'GRN in Money for senior roles, Work for site team',
    'Payments filtered (advance+final only) for Design/Services Head',
    'Detailing Head merged into Team Lead (same person, same nav)',
    'Pinned Needs You sections inside Work — all roles as shown in Sheet 4',
    'Today\'s Report pinned in Work for Site Manager / Senior Site Manager',
    'Pending tab (4th button) for Principal + Design Principal + Audit',
    'PMC Head: Needs You in Work + Blocked as separate Pending tab',
    'Design Head / Services Head: Needs You pinned in Work only (no separate Pending)',
    'Dual nudge: Day N WhatsApp to owner, Day N+2 WhatsApp to principal',
    'SLA defaults as shown in Sheet 6 (per-project, configurable)',
    'Drawings for Principal: visible inside Project Summary only (no separate tab)',
    'Drawings for PMC Head: in Work bucket (read-only)',
    'Project scoping bug: 7 field roles need to be scoped in routes/projects.js',
    'Daily report submission flow (new feature — site manager submits, PMC corrects, 6am auto-publish)',
    'Sort toggle on all lists: Urgency / Age / Default (role-specific default)',
    'Tab ordering within buckets: fixed by frequency now, adaptive later',
    'ROLE_TABS to be migrated from hardcoded JS to DB table (role_nav)',
    'project_slas table to be created for per-project SLA configuration',
    'Trainee + Detailing: 2-tab top strip, no bottom nav bar',
]

row = 5
for idx, item in enumerate(ITEMS, 1):
    cl(ws7, row, 1, idx, BG_SUB, F_BODY, C)
    cl(ws7, row, 2, item, BG_HIDE, F_SMALL, L)
    for col in (3, 4):
        c = ws7.cell(row=row, column=col)
        c.fill = BG_YELL; c.border = BOX; c.alignment = L
    ws7.row_dimensions[row].height = 30
    row += 1

ws7.column_dimensions['A'].width = 5
ws7.column_dimensions['B'].width = 52
ws7.column_dimensions['C'].width = 26
ws7.column_dimensions['D'].width = 40

# ─── Save ───
import os
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
OUT = '/mnt/user-data/outputs/20260420 nu PMC NavRedesign Audit v1.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}  |  Sheets: {len(wb.sheetnames)}')
