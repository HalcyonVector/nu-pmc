from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# ─── Styling ───
NAVY = '1D3D62'
BG_HEADER = PatternFill('solid', start_color=NAVY)
BG_HOME = PatternFill('solid', start_color='E3EFF9')
BG_WORK = PatternFill('solid', start_color='FFF3E0')
BG_MONEY = PatternFill('solid', start_color='E8F4EE')
BG_MORE = PatternFill('solid', start_color='F1ECE3')
BG_HIDDEN = PatternFill('solid', start_color='F8F8F8')
BG_YELLOW = PatternFill('solid', start_color='FFFBEF')

F_TITLE = Font(name='Arial', size=14, bold=True, color=NAVY)
F_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY = Font(name='Arial', size=10)
F_BODY_BOLD = Font(name='Arial', size=10, bold=True, color=NAVY)
F_BUCKET = Font(name='Arial', size=11, bold=True, color=NAVY)
F_MUTED = Font(name='Arial', size=9, color='888888', italic=True)

thin = Side(border_style='thin', color='CCCCCC')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ═══════════════════════════════════════════════
# ROLE → BUCKET → TABS
# Every tab each role currently has, mapped to Home / Work / Money / More
# ═══════════════════════════════════════════════

# Tab label lookup
TAB_LABEL = {
    'dashboard':'Dashboard', 'monthly':'Monthly Overview', 'projects':'Projects',
    'project_detail':'Project Summary', 'approvals':'Approvals', 'signoff':'Weekly Sign-off',
    'changes':'CNs', 'delegations':'Delegations', 'budget':'Budget',
    'budget_tree':'Budget Tree', 'vendors':'Engagements', 'vendors_master':'Vendor Master',
    'boq_mapping':'BOQ Map', 'materials':'Materials', 'materials_site':'Materials (site)',
    'client_boq':'Client BOQ', 'payments':'Payments', 'payments_fin':'Payments (Finance)',
    'finance_clearance':'Vendor Clearance', 'weekly_health':'Health', 'users':'Users',
    'drawings':'Drawings', 'register':'Register', 'submittals':'Submittals',
    'grn':'GRNs', 'issues':'Issues', 'issues_site':'Field Issues',
    'labour':'Labour', 'photos':'Photos', 'phototags':'Photo Review',
    'tasks':'Tasks', 'schedule_view':'Schedule', 'meetings':'Meetings',
    'reports':'Reports', 'deputy':'Deputy', 'clients':'Clients',
    'pi':'Invoices (PI)', 'petty_cash':'Petty Cash', 'client_receipts':'Receipts',
    'gst_statement':'GST', 'ncr':'NCRs', 'compliance':'Compliance',
    'tally':'Tally', 'notifications':'Alerts',
}

# Per-role bucket mapping
# Each role → dict of bucket → ordered list of tab keys
ROLE_BUCKETS = {
    'Principal': {
        'Home':  ['dashboard', 'monthly', 'projects', 'project_detail'],
        'Work':  ['weekly_health'],
        'Money': ['payments', 'payments_fin', 'vendors', 'vendors_master', 'finance_clearance', 'budget', 'budget_tree', 'materials', 'boq_mapping', 'client_boq'],
        'More':  ['approvals', 'signoff', 'changes', 'delegations', 'users'],
    },
    'Design Principal': {
        'Home':  ['dashboard', 'monthly', 'projects', 'project_detail'],
        'Work':  ['weekly_health'],
        'Money': ['payments', 'payments_fin', 'vendors', 'vendors_master', 'finance_clearance', 'budget', 'materials', 'boq_mapping', 'client_boq'],
        'More':  ['approvals', 'signoff', 'changes', 'delegations', 'users'],
    },
    'PMC Head': {
        'Home':  ['dashboard', 'monthly', 'project_detail'],
        'Work':  ['reports', 'grn', 'issues', 'meetings', 'labour'],
        'Money': ['payments', 'vendors', 'vendors_master', 'boq_mapping', 'materials', 'client_boq'],
        'More':  ['signoff', 'delegations', 'deputy'],
    },
    'Design Head': {
        'Home':  ['dashboard', 'monthly', 'project_detail'],
        'Work':  ['drawings', 'register', 'submittals', 'issues', 'phototags'],
        'Money': ['payments', 'vendors', 'vendors_master', 'boq_mapping', 'materials', 'client_boq', 'budget'],
        'More':  ['signoff', 'delegations'],
    },
    'Services Head': {
        'Home':  ['dashboard', 'monthly', 'project_detail'],
        'Work':  ['drawings', 'register', 'submittals', 'issues', 'phototags'],
        'Money': ['payments', 'vendors', 'vendors_master', 'boq_mapping', 'materials', 'client_boq', 'budget'],
        'More':  ['signoff', 'delegations'],
    },
    'Site Manager': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['tasks', 'drawings', 'register', 'issues', 'issues_site', 'photos', 'labour', 'materials_site'],
        'Money': ['grn', 'payments'],
        'More':  [],
    },
    'Senior Site Manager': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['tasks', 'drawings', 'register', 'issues', 'issues_site', 'photos', 'labour', 'materials_site'],
        'Money': ['grn', 'payments'],
        'More':  [],
    },
    'Finance Admin': {
        'Home':  ['dashboard', 'monthly', 'project_detail'],
        'Work':  [],
        'Money': ['payments_fin', 'finance_clearance', 'vendors_master', 'clients', 'client_boq', 'pi', 'petty_cash', 'client_receipts', 'gst_statement'],
        'More':  [],
    },
    'Trainee': {
        'Home':  [],
        'Work':  ['drawings', 'schedule_view'],
        'Money': [],
        'More':  [],
    },
    'Team Lead': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['drawings', 'issues', 'submittals'],
        'Money': [],
        'More':  [],
    },
    'Detailing Head': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['drawings', 'register', 'issues', 'submittals', 'phototags'],
        'Money': [],
        'More':  [],
    },
    'Jr Architect': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['drawings', 'issues', 'submittals'],
        'Money': [],
        'More':  [],
    },
    'Detailing': {
        'Home':  [],
        'Work':  ['drawings', 'submittals'],
        'Money': [],
        'More':  [],
    },
    'Services Engineer': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['drawings', 'issues', 'submittals', 'phototags'],
        'Money': [],
        'More':  [],
    },
    'Coordinator': {
        'Home':  ['dashboard', 'project_detail'],
        'Work':  ['drawings', 'register', 'issues', 'tasks', 'photos', 'meetings'],
        'Money': ['grn'],
        'More':  [],
    },
    'Audit': {
        'Home':  ['dashboard', 'monthly', 'projects', 'project_detail'],
        'Work':  ['drawings', 'register', 'submittals', 'issues', 'issues_site', 'meetings', 'reports', 'tasks', 'photos', 'phototags', 'labour', 'schedule_view'],
        'Money': ['payments', 'payments_fin', 'vendors', 'vendors_master', 'finance_clearance', 'clients', 'client_boq', 'boq_mapping', 'materials', 'budget', 'budget_tree', 'pi', 'petty_cash', 'client_receipts', 'gst_statement', 'grn'],
        'More':  ['approvals', 'signoff', 'changes', 'delegations', 'deputy', 'users', 'weekly_health', 'ncr', 'compliance', 'tally', 'notifications'],
    },
}

# ═══════════════════════════════════════════════
# SHEET 1 — Summary & Rules
# ═══════════════════════════════════════════════
ws = wb.create_sheet('Read me first')
ws['A1'] = 'nu PMC — Navigation Redesign Audit'
ws['A1'].font = F_TITLE
ws.merge_cells('A1:E1')
ws['A2'] = '4-bucket bottom navigation: Home / Work / Money / More'
ws['A2'].font = F_MUTED
ws.merge_cells('A2:E2')

# Rules section
rules = [
    ('Rule', 'What it means'),
    ('4 buckets max', 'Every role sees at most Home / Work / Money / More at the bottom. Empty buckets hide.'),
    ('Home', 'Where you land. Dashboard + where-am-I-in-the-firm views (Monthly / Projects) + Project Summary.'),
    ('Work', 'Role-specific daily verbs. What the person actually does every day on this app.'),
    ('Money', 'Anything money-adjacent. Payments, vendors, GRN, BOQ, budget, materials approval, invoicing.'),
    ('More', 'Governance, admin, rarer actions. Approvals, signoff, delegations, users, CNs.'),
    ('Bucket content changes per role', 'A tab like Drawings is "Work" for Design Head (they manage it) but also "Work" for Site Manager (they consume it). GRN is "Work" for Site Manager (raise) but "Money" for Principal (approve).'),
    ('Empty bucket is fine', 'Site Manager has no "More". Finance has no "Work". Bar renders 2-4 buttons as needed.'),
    ('Tab count per role', 'Does not change from current spec. Only grouping changes.'),
]

row = 4
for i, (col_a, col_b) in enumerate(rules):
    if i == 0:
        c1 = ws.cell(row=row, column=1, value=col_a); c1.font = F_HEADER; c1.fill = BG_HEADER; c1.alignment = CENTER; c1.border = BOX
        c2 = ws.cell(row=row, column=2, value=col_b); c2.font = F_HEADER; c2.fill = BG_HEADER; c2.alignment = CENTER; c2.border = BOX
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    else:
        c1 = ws.cell(row=row, column=1, value=col_a); c1.font = F_BODY_BOLD; c1.alignment = LEFT_CENTER; c1.border = BOX; c1.fill = PatternFill('solid', start_color='E8EEF4')
        c2 = ws.cell(row=row, column=2, value=col_b); c2.font = F_BODY; c2.alignment = LEFT; c2.border = BOX
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 34
    row += 1

ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20

# Summary of per-role bucket counts
row += 2
c = ws.cell(row=row, column=1, value='Per-role bucket summary')
c.font = F_BUCKET
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

summary_headers = ['Role', 'Home', 'Work', 'Money', 'More']
for col_idx, h in enumerate(summary_headers, 1):
    c = ws.cell(row=row, column=col_idx, value=h)
    c.font = F_HEADER; c.fill = BG_HEADER; c.alignment = CENTER; c.border = BOX
row += 1

for role, buckets in ROLE_BUCKETS.items():
    c = ws.cell(row=row, column=1, value=role)
    c.font = F_BODY_BOLD; c.border = BOX
    c.fill = PatternFill('solid', start_color='F4F2ED')
    c.alignment = LEFT_CENTER
    for col_idx, bucket in enumerate(['Home','Work','Money','More'], 2):
        n = len(buckets.get(bucket, []))
        c = ws.cell(row=row, column=col_idx, value=n if n > 0 else '—')
        c.alignment = CENTER
        c.border = BOX
        fill = {'Home': BG_HOME, 'Work': BG_WORK, 'Money': BG_MONEY, 'More': BG_MORE}[bucket]
        c.fill = fill if n > 0 else BG_HIDDEN
        c.font = F_BODY
    row += 1

# ═══════════════════════════════════════════════
# SHEET 2 — Per-role sheets, compact format
# One sheet per role. 4 rows (Home/Work/Money/More) × content
# ═══════════════════════════════════════════════

def fmt_tab(key):
    return TAB_LABEL.get(key, key)

for role, buckets in ROLE_BUCKETS.items():
    sheet_name = role[:31]
    ws = wb.create_sheet(sheet_name)

    ws['A1'] = f'{role} — bottom nav layout'
    ws['A1'].font = F_TITLE
    ws.merge_cells('A1:D1')
    n_visible = sum(1 for b in ['Home','Work','Money','More'] if buckets.get(b))
    ws['A2'] = f'{n_visible} bottom-bar button{"s" if n_visible != 1 else ""} visible. Empty buckets hide.'
    ws['A2'].font = F_MUTED
    ws.merge_cells('A2:D2')

    # Header row: 4 bucket headers
    headers = ['Home', 'Work', 'Money', 'More']
    fills = [BG_HOME, BG_WORK, BG_MONEY, BG_MORE]
    for col_idx, (h, fill) in enumerate(zip(headers, fills), 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = F_HEADER
        c.fill = BG_HEADER
        c.alignment = CENTER
        c.border = BOX
    ws.row_dimensions[4].height = 28

    # Content row — each bucket's tab list
    for col_idx, (bucket, fill) in enumerate(zip(headers, fills), 1):
        tabs = buckets.get(bucket, [])
        if not tabs:
            c = ws.cell(row=5, column=col_idx, value='(hidden — no tabs in this bucket)')
            c.font = F_MUTED
            c.fill = BG_HIDDEN
        else:
            body = '\n'.join(f'• {fmt_tab(t)}' for t in tabs)
            c = ws.cell(row=5, column=col_idx, value=body)
            c.font = F_BODY
            c.fill = fill
        c.alignment = LEFT
        c.border = BOX
    # Height based on max bucket size
    max_tabs = max((len(buckets.get(b, [])) for b in headers), default=0)
    ws.row_dimensions[5].height = max(60, 22 + 18 * max_tabs)

    # Column widths — wider for Work/Money since they have more content
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 22

    # Audit note row
    c = ws.cell(row=7, column=1, value='Your audit notes (changes to make):')
    c.font = F_BODY_BOLD
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=4)
    for col_idx, fill in enumerate(fills, 1):
        c = ws.cell(row=8, column=col_idx, value='')
        c.fill = BG_YELLOW
        c.border = BOX
        c.alignment = LEFT
    ws.row_dimensions[8].height = 80

# Save
import os
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
OUT = '/mnt/user-data/outputs/20260420 nu BottomNav Audit v1.xlsx'
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheets: {len(wb.sheetnames)}')
